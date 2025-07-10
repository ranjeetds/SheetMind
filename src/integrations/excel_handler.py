"""
Excel integration handler for SheetMind.

Provides unified interface for Excel operations across different platforms
and integration methods (COM API, file-based, etc.).
"""

import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Platform-specific imports
if sys.platform == "win32":
    try:
        import win32com.client as win32
        HAS_WIN32_COM = True
    except ImportError:
        HAS_WIN32_COM = False
else:
    HAS_WIN32_COM = False


class ExcelHandler:
    """
    Unified Excel handler supporting multiple integration methods.
    
    Automatically chooses the best available method based on platform and requirements.
    """
    
    def __init__(self, file_path: Optional[str] = None, use_com: bool = None):
        """
        Initialize Excel handler.
        
        Args:
            file_path: Path to Excel file to work with
            use_com: Force COM API usage (Windows only), None for auto-detect
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.data = None
        
        # Determine integration method
        if use_com is None:
            self.use_com = HAS_WIN32_COM and sys.platform == "win32"
        else:
            self.use_com = use_com and HAS_WIN32_COM
        
        # COM API objects
        self.excel_app = None
        self.com_workbook = None
        self.com_worksheet = None
        
        if file_path:
            self.load_file(file_path)
    
    def load_file(self, file_path: str) -> bool:
        """
        Load an Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            self.file_path = file_path
            
            if self.use_com:
                return self._load_file_com(file_path)
            else:
                return self._load_file_openpyxl(file_path)
                
        except Exception as e:
            print(f"Error loading file: {e}")
            return False
    
    def _load_file_com(self, file_path: str) -> bool:
        """Load file using COM API."""
        try:
            # Initialize Excel application
            self.excel_app = win32.Dispatch("Excel.Application")
            self.excel_app.Visible = False
            
            # Open workbook
            abs_path = os.path.abspath(file_path)
            self.com_workbook = self.excel_app.Workbooks.Open(abs_path)
            self.com_worksheet = self.com_workbook.ActiveSheet
            
            return True
        except Exception as e:
            print(f"COM API error: {e}")
            return False
    
    def _load_file_openpyxl(self, file_path: str) -> bool:
        """Load file using openpyxl."""
        try:
            if os.path.exists(file_path):
                self.workbook = load_workbook(file_path)
                self.worksheet = self.workbook.active
            else:
                # Create new workbook
                self.workbook = Workbook()
                self.worksheet = self.workbook.active
            
            # Load data as DataFrame for easier manipulation
            self.data = pd.read_excel(file_path) if os.path.exists(file_path) else pd.DataFrame()
            
            return True
        except Exception as e:
            print(f"Openpyxl error: {e}")
            return False
    
    def get_data(self, sheet_name: str = None) -> pd.DataFrame:
        """
        Get data from the current worksheet as a DataFrame.
        
        Args:
            sheet_name: Optional sheet name, uses active sheet if None
            
        Returns:
            DataFrame containing the data
        """
        if self.use_com and self.com_worksheet:
            return self._get_data_com(sheet_name)
        elif self.workbook:
            return self._get_data_openpyxl(sheet_name)
        else:
            return pd.DataFrame()
    
    def _get_data_com(self, sheet_name: str = None) -> pd.DataFrame:
        """Get data using COM API."""
        try:
            if sheet_name:
                worksheet = self.com_workbook.Sheets(sheet_name)
            else:
                worksheet = self.com_worksheet
            
            # Get used range
            used_range = worksheet.UsedRange
            if used_range is None:
                return pd.DataFrame()
            
            # Convert to DataFrame
            data = list(used_range.Value)
            if data and data[0]:
                df = pd.DataFrame(data[1:], columns=data[0])
                return df
            
            return pd.DataFrame()
        except Exception as e:
            print(f"Error getting COM data: {e}")
            return pd.DataFrame()
    
    def _get_data_openpyxl(self, sheet_name: str = None) -> pd.DataFrame:
        """Get data using openpyxl."""
        try:
            if sheet_name and sheet_name in self.workbook.sheetnames:
                ws = self.workbook[sheet_name]
            else:
                ws = self.worksheet
            
            # Convert to DataFrame
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            
            if data:
                df = pd.DataFrame(data[1:], columns=data[0])
                return df
            
            return pd.DataFrame()
        except Exception as e:
            print(f"Error getting openpyxl data: {e}")
            return pd.DataFrame()
    
    def write_data(self, data: pd.DataFrame, sheet_name: str = None, start_row: int = 1, start_col: int = 1):
        """
        Write DataFrame to worksheet.
        
        Args:
            data: DataFrame to write
            sheet_name: Target sheet name
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)
        """
        if self.use_com:
            self._write_data_com(data, sheet_name, start_row, start_col)
        else:
            self._write_data_openpyxl(data, sheet_name, start_row, start_col)
    
    def _write_data_com(self, data: pd.DataFrame, sheet_name: str = None, start_row: int = 1, start_col: int = 1):
        """Write data using COM API."""
        try:
            if sheet_name:
                worksheet = self.com_workbook.Sheets(sheet_name)
            else:
                worksheet = self.com_worksheet
            
            # Write headers
            for i, col in enumerate(data.columns):
                worksheet.Cells(start_row, start_col + i).Value = col
            
            # Write data
            for i, row in data.iterrows():
                for j, value in enumerate(row):
                    worksheet.Cells(start_row + i + 1, start_col + j).Value = value
                    
        except Exception as e:
            print(f"Error writing COM data: {e}")
    
    def _write_data_openpyxl(self, data: pd.DataFrame, sheet_name: str = None, start_row: int = 1, start_col: int = 1):
        """Write data using openpyxl."""
        try:
            if sheet_name:
                if sheet_name in self.workbook.sheetnames:
                    ws = self.workbook[sheet_name]
                else:
                    ws = self.workbook.create_sheet(sheet_name)
            else:
                ws = self.worksheet
            
            # Clear existing data
            ws.delete_rows(start_row, ws.max_row)
            
            # Write data
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)
                
        except Exception as e:
            print(f"Error writing openpyxl data: {e}")
    
    def add_formula(self, formula: str, cell: str, sheet_name: str = None) -> bool:
        """
        Add a formula to a specific cell.
        
        Args:
            formula: Excel formula (e.g., "=SUM(A1:A10)")
            cell: Target cell (e.g., "B1")
            sheet_name: Target sheet name
            
        Returns:
            True if successful
        """
        try:
            if self.use_com:
                return self._add_formula_com(formula, cell, sheet_name)
            else:
                return self._add_formula_openpyxl(formula, cell, sheet_name)
        except Exception as e:
            print(f"Error adding formula: {e}")
            return False
    
    def _add_formula_com(self, formula: str, cell: str, sheet_name: str = None) -> bool:
        """Add formula using COM API."""
        try:
            if sheet_name:
                worksheet = self.com_workbook.Sheets(sheet_name)
            else:
                worksheet = self.com_worksheet
            
            worksheet.Range(cell).Formula = formula
            return True
        except Exception as e:
            print(f"COM formula error: {e}")
            return False
    
    def _add_formula_openpyxl(self, formula: str, cell: str, sheet_name: str = None) -> bool:
        """Add formula using openpyxl."""
        try:
            if sheet_name and sheet_name in self.workbook.sheetnames:
                ws = self.workbook[sheet_name]
            else:
                ws = self.worksheet
            
            ws[cell] = formula
            return True
        except Exception as e:
            print(f"Openpyxl formula error: {e}")
            return False
    
    def create_chart(self, chart_type: str, data_range: str, title: str = "", sheet_name: str = None) -> bool:
        """
        Create a chart from data range.
        
        Args:
            chart_type: Type of chart ("bar", "line", "pie")
            data_range: Data range (e.g., "A1:C10")
            title: Chart title
            sheet_name: Target sheet name
            
        Returns:
            True if successful
        """
        try:
            if self.use_com:
                return self._create_chart_com(chart_type, data_range, title, sheet_name)
            else:
                return self._create_chart_openpyxl(chart_type, data_range, title, sheet_name)
        except Exception as e:
            print(f"Error creating chart: {e}")
            return False
    
    def _create_chart_openpyxl(self, chart_type: str, data_range: str, title: str = "", sheet_name: str = None) -> bool:
        """Create chart using openpyxl."""
        try:
            if sheet_name and sheet_name in self.workbook.sheetnames:
                ws = self.workbook[sheet_name]
            else:
                ws = self.worksheet
            
            # Parse data range
            data = Reference(ws, range_string=data_range)
            
            # Create chart based on type
            if chart_type.lower() == "bar":
                chart = BarChart()
            elif chart_type.lower() == "line":
                chart = LineChart()
            elif chart_type.lower() == "pie":
                chart = PieChart()
            else:
                chart = BarChart()  # Default
            
            chart.add_data(data, titles_from_data=True)
            chart.title = title
            
            # Add chart to worksheet
            ws.add_chart(chart, "E5")
            return True
            
        except Exception as e:
            print(f"Openpyxl chart error: {e}")
            return False
    
    def save(self, file_path: str = None) -> bool:
        """
        Save the workbook.
        
        Args:
            file_path: Optional new file path
            
        Returns:
            True if successful
        """
        try:
            save_path = file_path or self.file_path
            
            if self.use_com and self.com_workbook:
                if file_path:
                    self.com_workbook.SaveAs(os.path.abspath(file_path))
                else:
                    self.com_workbook.Save()
                return True
            elif self.workbook:
                self.workbook.save(save_path)
                return True
            
            return False
        except Exception as e:
            print(f"Error saving file: {e}")
            return False
    
    def close(self):
        """Close the Excel application and workbook."""
        try:
            if self.use_com:
                if self.com_workbook:
                    self.com_workbook.Close()
                if self.excel_app:
                    self.excel_app.Quit()
        except Exception as e:
            print(f"Error closing Excel: {e}")
    
    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names."""
        try:
            if self.use_com and self.com_workbook:
                return [sheet.Name for sheet in self.com_workbook.Sheets]
            elif self.workbook:
                return self.workbook.sheetnames
            else:
                return []
        except Exception as e:
            print(f"Error getting sheet names: {e}")
            return []
    
    def create_sheet(self, name: str) -> bool:
        """Create a new worksheet."""
        try:
            if self.use_com and self.com_workbook:
                self.com_workbook.Sheets.Add().Name = name
                return True
            elif self.workbook:
                self.workbook.create_sheet(name)
                return True
            else:
                return False
        except Exception as e:
            print(f"Error creating sheet: {e}")
            return False 